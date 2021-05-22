import argparse
import json
import logging
import os
import sqlite3
import sys
from pathlib import Path

import openpyxl as opx


def get_logger(logger_file, name=__file__, encoding="utf-8"):
    log = logging.getLogger(name)
    log.setLevel(logging.DEBUG)
    formatter = logging.Formatter(
        "[%(asctime)s] %(filename)s:%(lineno)d %(levelname)-8s %(message)s"
    )
    fh = logging.FileHandler(logger_file, encoding=encoding)
    fh.setFormatter(formatter)
    log.addHandler(fh)
    return log


class Unit:
    def __init__(self, name: str, unit_id):
        self.name = name
        self.unit_id = unit_id
        self.__load_max = None
        self.input_stream = {}
        self.output_stream = {}

    def set_load_max(self, load_max_value):
        self.__load_max = load_max_value

    @property
    def load_max(self):
        return self.__load_max

    def __repr__(self):
        return f"Unit_{self.name}"


class ABTUnit(Unit):
    pass


class SecondaryUnit(Unit):
    pass


class Stream:
    def __init__(self, name: str, stream_id: int):
        self.name = name
        self.id = stream_id
        self.where_from = []
        self.where_to = []

    def add_where_from(self, unit: Unit):
        self.where_from.append(unit)

    def add_where_to(self, unit: Unit):
        self.where_to.append(unit)

    def __repr__(self):
        return f"Stream_{self.name}"


def save_cursor_executor_creator(cursor, my_logger):
    def wrapper(command):
        try:
            cursor.execute(command)
        except Exception as e:
            my_logger.critical(
                f"Some problems with DB: '{e}'. Program stopped!\n"
                f"--------------------------------------------------------------------------------"
            )
            sys.exit(1)

    return wrapper


if __name__ == "__main__":

    DEFAULT_DATABASE_NAME = "db.db"
    DEFAULT_JSON_FILENAME = "multiple_streams.json"
    DEFAULT_CSV_FILENAME = "unused_streams.csv"
    DEFAULT_XLSX_FILENAME = "all_units.xlsx"
    LOGGER_FILENAME = "log.txt"

    my_logger = get_logger(logger_file=LOGGER_FILENAME)

    my_logger.info(" Program started")

    parser = argparse.ArgumentParser(description="Units and Streams analyser")
    parser.add_argument(
        "-db_file",
        type=str,
        help="Input database file",
        default=DEFAULT_DATABASE_NAME,
    )
    parser.add_argument(
        "-json_filename",
        type=str,
        help="Name of the output .json file with multiple streams",
        default=DEFAULT_JSON_FILENAME,
    )
    parser.add_argument(
        "-csv_filename",
        type=str,
        help="Name of the output .csv file with unused streams",
        default=DEFAULT_CSV_FILENAME,
    )
    parser.add_argument(
        "-xlsx_filename",
        type=str,
        help="Name of the output .xlsx file with unit streams",
        default=DEFAULT_XLSX_FILENAME,
    )

    args = parser.parse_args()
    db_file = args.db_file

    if not args.json_filename.endswith(".json"):
        json_filename = Path(args.json_filename + ".json")
    else:
        json_filename = Path(args.json_filename)

    if not args.csv_filename.endswith(".csv"):
        csv_filename = Path(args.csv_filename + ".csv")
    else:
        csv_filename = Path(args.csv_filename)

    if not args.xlsx_filename.endswith(".xlsx"):
        xlsx_filename = Path(args.xlsx_filename + ".xlsx")
    else:
        xlsx_filename = Path(args.xlsx_filename)

    if not os.path.exists(db_file):
        my_logger.critical("Database doesn't exists. Program stopped!")
        sys.exit(1)

    my_logger.info(f"Connecting to database '{db_file}'")

    try:
        con = sqlite3.connect(db_file)
        cursor = con.cursor()
    except Exception as e:
        my_logger.critical("Problems with database connection. Program stopped!")
        sys.exit(1)

    save_cursor_execute = save_cursor_executor_creator(cursor, my_logger)

    # Задание 1
    my_logger.info(f"Creating unit items (task 1)")
    save_cursor_execute("SELECT * FROM unit")
    units = {
        unit_id: SecondaryUnit(name, unit_id) if unit_type else ABTUnit(name, unit_id)
        for unit_id, name, unit_type in cursor
    }

    # Задание 2
    my_logger.info(f"Creating stream items (task 2)")
    save_cursor_execute("SELECT * FROM stream")
    streams = {stream_id: Stream(name, stream_id) for stream_id, name in cursor}

    # Задание 3
    my_logger.info(f"Querying the database with a big request (task 3)")
    save_cursor_execute(
        "SELECT unit.name, stream.name  FROM unit, stream, unit_material "
        "WHERE unit_material.unit_id = unit.id AND unit_material.stream_id = stream.id "
        "AND unit_material.feed_flag = 1 "
    )
    # В задании не сказано, что делать с полученным ответом на запрос, поэтому я его просто вывел
    sys.stdout.write(str(cursor.fetchall()))

    my_logger.info(f"Connecting units and streams")
    save_cursor_execute("SELECT * FROM unit_material")
    for unit_id, stream_id, feed_flag in cursor:
        if feed_flag:
            units[unit_id].input_stream[stream_id] = streams[stream_id]
            streams[stream_id].add_where_to(units[unit_id])
        else:
            units[unit_id].output_stream[stream_id] = streams[stream_id]
            streams[stream_id].add_where_from(units[unit_id])

    save_cursor_execute("SELECT * FROM load_max")
    for unit_id, load_max in cursor:
        units[unit_id].set_load_max(load_max)

    cursor.close()

    # Задание 4
    my_logger.info(f"Writing {csv_filename} file (task 4)")
    try:
        with open(csv_filename, "w") as f:
            for stream in streams.values():
                if not stream.where_from and not stream.where_to:
                    f.write(f"{stream.id}, {stream.name}\n")
    except PermissionError as e:
        my_logger.warning(f"Impossible to save '{csv_filename}'. {e}")

    # Задание 5
    my_logger.info(f"Writing {json_filename} file (task 5)")
    try:
        with open(json_filename, "w") as file:
            json.dump(
                {
                    stream.name: [unit.name for unit in stream.where_to]
                    for stream in streams.values()
                    if len(stream.where_to) > 1
                },
                file,
                indent=4,
            )
    except PermissionError as e:
        my_logger.warning(f"Impossible to save '{json_filename}'. {e}")

    # Задание 6
    my_logger.info(f"Writing {xlsx_filename} file (task 6)")
    wb = opx.Workbook()
    wb.remove(wb.active)
    for unit in units.values():
        ws = wb.create_sheet(unit.name)
        for index, input_stream in enumerate(unit.input_stream.values(), 1):
            ws[f"A{index}"] = input_stream.name
        for index, output_stream in enumerate(unit.output_stream.values(), 1):
            ws[f"B{index}"] = output_stream.name

    try:
        wb.save(xlsx_filename)
    except PermissionError as e:
        my_logger.warning(f"Impossible to save '{xlsx_filename}'. {e}")

    my_logger.info(
        "Program finished\n"
        "--------------------------------------------------------------------------------"
    )
